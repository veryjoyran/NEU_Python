import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook
from sklearn.linear_model import LinearRegression
import logging

# 配置日志
logging.basicConfig(
    filename='system.log',  # 日志文件名
    level=logging.INFO,     # 设置日志记录级别为 INFO
    format='%(asctime)s - %(levelname)s - %(message)s'  # 日志格式
)

file_path = 'house_data_20240430.xlsx'

def load_data_with_progress(file_path):
    workbook = load_workbook(filename=file_path, read_only=True)
    sheet = workbook.active

    # 获取总行数，减去标题行
    total_rows = sheet.max_row - 1

    # 使用 tqdm 显示进度条
    data = []
    header = None
    for i, row in enumerate(tqdm(sheet.iter_rows(values_only=True), total=total_rows + 1)):
        if i == 0:
            header = row  # 第一行作为标题
        else:
            data.append(row)  # 剩余行作为数据内容

    df = pd.DataFrame(data, columns=header)
    return df

house_data = load_data_with_progress(file_path)
logging.info("数据加载完成，共加载 %d 条记录", len(house_data))

column_translation = {
    'link': '链接',
    'buildingCount': '楼栋数量',
    'buildingType': '建筑类型',
    'electricType': '电力类型',
    'favCount': '收藏数量',
    'greenRate': '绿化率',
    'heatingType': '供暖类型',
    'name': '名称',
    'property': '物业类型',
    'resblockId': '小区ID',
    'shopDistance': '商圈距离',
    'unitPrice': '单价',
    'waterType': '水源类型',
    'f地址': '地址',
    '商圈': '商圈',
    '省份': '省份',
    '城市': '城市',
    '区县': '区县',
    '街镇': '街镇'
}

house_data.rename(columns=column_translation, inplace=True)

pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
pd.set_option('display.width', 1000)
pd.set_option('display.colheader_justify', 'left')

query_address = input("请输入查询地址关键词: ")
logging.info("用户输入的查询关键词: %s", query_address)

filtered_data = house_data[house_data['地址'].str.contains(query_address, na=False)]
logging.info("筛选出 %d 条符合条件的记录", len(filtered_data))

# 填充房价空值
if filtered_data['单价'].isna().any():
    print("查询结果中存在房价空值，选择填充方式：")
    print("1. 使用均值填充")
    print("2. 使用线性回归预测填充")
    fill_option = input("请输入填充方式编号（1/2）：")

    filtered_data.loc[:, '单价'] = filtered_data['单价'].astype(str)

    if fill_option == '1':
        mean_price = filtered_data['单价'].astype(float).mean()
        filtered_data.loc[filtered_data['单价'] == 'nan', '单价'] = f"{mean_price:.2f}（预测值）"
        logging.info("使用均值填充房价空值，填充值为 %.2f", mean_price)
        print("已使用均值填充房价空值")

    elif fill_option == '2':
        features = filtered_data[['收藏数量', '楼栋数量']].copy()  # 示例特征
        features['收藏数量'] = features['收藏数量'].fillna(features['收藏数量'].mean())
        features['楼栋数量'] = features['楼栋数量'].fillna(features['楼栋数量'].mean())

        known_price = filtered_data[filtered_data['单价'] != 'nan']
        unknown_price = filtered_data[filtered_data['单价'] == 'nan']

        model = LinearRegression()
        model.fit(known_price[['收藏数量', '楼栋数量']], known_price['单价'].astype(float))

        predicted_prices = model.predict(unknown_price[['收藏数量', '楼栋数量']])
        filtered_data.loc[filtered_data['单价'] == 'nan', '单价'] = [f"{price:.2f}（预测值）" for price in predicted_prices]
        logging.info("使用线性回归预测填充房价空值，共填充 %d 条记录", len(predicted_prices))
        print("已使用线性回归预测填充房价空值")

    else:
        logging.warning("无效的填充选项，未进行填充")
        print("无效选择，未进行填充")
else:
    logging.info("查询结果中无房价空值")
    print("查询结果中无房价空值")

print("请选择排序方式：")
print("1. 按收藏数量降序排序")
print("2. 按价格升序排序")
print("3. 按价格降序排序")
sort_option = input("请输入排序方式的编号（1/2/3）：")

if sort_option == '1':
    sorted_data = filtered_data.sort_values(by='收藏数量', ascending=False)
    logging.info("按收藏数量降序排序完成")
elif sort_option == '2':
    sorted_data = filtered_data.sort_values(by='单价', ascending=True, key=lambda x: pd.to_numeric(x.str.replace('（预测值）', ''), errors='coerce'))
    logging.info("按价格升序排序完成")
elif sort_option == '3':
    sorted_data = filtered_data.sort_values(by='单价', ascending=False, key=lambda x: pd.to_numeric(x.str.replace('（预测值）', ''), errors='coerce'))
    logging.info("按价格降序排序完成")
else:
    sorted_data = filtered_data.sort_values(by='收藏数量', ascending=False)
    logging.warning("无效的排序选项，默认按收藏数量降序排序")
    print("无效的输入，默认按收藏数量降序排序")

output_columns = [
    '单价', '收藏数量', '地址', '商圈', '商圈距离', '水源类型', '电力类型',
    '楼栋数量', '绿化率', '物业类型', '小区ID', '链接'
]

result = sorted_data[output_columns]
print(result)
result.to_excel("filtered_results.xlsx", index=False)
logging.info("查询结果已保存到文件: filtered_results.xlsx")
